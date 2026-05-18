"""
Build a Content Library Contradictions Excel report.

Usage:
  python build_audit_xlsx.py --input findings.json --output report.xlsx

findings.json schema (see references/output_format.md for full spec):
{
  "scope": "Whole library | Security | Integrations | ...",
  "library_size": 11102,
  "items_audited": 4250,
  "ground_truth_anchors": [
    {"fact": "...", "current_truth": "...", "source": "user-supplied | derived-from-recent"}
  ],
  "conflicting_facts": [
    {
      "topic": "Languages supported",
      "item_a": {"question": "...", "answer_excerpt": "...", "file": "...", "date": "YYYY-MM-DD",
                 "usage": 25, "reference_url": "https://app.autorfp.ai/references?contentId=..."},
      "item_b": {...},
      "conflict": "Item A says 25+ languages, Item B says 44+ languages",
      "recommendation": "Keep Item B; archive or update Item A",
      "risk": "Critical | High | Medium | Verify"
    }
  ],
  "outdated_facts": [
    {
      "topic": "Regional deployments",
      "item": {"question": "...", "answer_excerpt": "...", "file": "...", "date": "YYYY-MM-DD",
               "usage": 17, "reference_url": "..."},
      "anchor_fact": "Multi-region (US, EU, AU)",
      "stale_claim": "All data hosted in Australia",
      "recommendation": "Update or archive",
      "risk": "Critical"
    }
  ],
  "capability_contradictions": [
    {
      "capability": "SharePoint integration",
      "negative_item": {"question": "...", "answer_excerpt": "...", "file": "...", "date": "...",
                        "usage": 8, "reference_url": "..."},
      "positive_item": {...},
      "recommendation": "Keep positive item; archive negative",
      "risk": "High"
    }
  ]
}
"""
import argparse
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------- Default palette (neutral, professional) ----------
NAVY = "0F1E3D"
ORANGE = "E26F2C"
LIGHT_GREY = "F4F4F4"
RISK_COLORS = {
    "Critical": "C0392B",
    "High": "E67E22",
    "Medium": "F1C40F",
    "Verify": "BDC3C7",
}

# ---------- Common styles ----------
TITLE_FONT = Font(name="Arial", size=18, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="Arial", size=11, color="FFFFFF")
H1_FONT = Font(name="Arial", size=14, bold=True, color=NAVY)
H2_FONT = Font(name="Arial", size=11, bold=True, color=NAVY)
HDR_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10)
BODY_BOLD = Font(name="Arial", size=10, bold=True)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
NAVY_FILL = PatternFill("solid", fgColor=NAVY)
ORANGE_FILL = PatternFill("solid", fgColor=ORANGE)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def add_header_band(ws, scope: str):
    """Branded title band at the top of every sheet."""
    ws.merge_cells("A1:H3")
    cell = ws["A1"]
    cell.value = "Content Library — Contradiction Audit"
    cell.font = TITLE_FONT
    cell.fill = NAVY_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    ws.merge_cells("A4:H4")
    sub = ws["A4"]
    sub.value = f"Scope: {scope}"
    sub.font = SUBTITLE_FONT
    sub.fill = ORANGE_FILL
    sub.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[4].height = 22


def risk_cell(ws, row, col, risk):
    """Write a coloured risk badge cell."""
    c = ws.cell(row=row, column=col, value=risk)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=RISK_COLORS.get(risk, "BDC3C7"))
    c.alignment = CENTER


def write_link(ws, row, col, url, label="Open in AutoRFP.ai"):
    """Write a clickable hyperlink cell. URLs are mandatory in every flagged row."""
    if not url:
        c = ws.cell(row=row, column=col, value="(no URL)")
        c.font = BODY
        return
    c = ws.cell(row=row, column=col, value=label)
    c.hyperlink = url
    c.font = LINK_FONT


def write_wrapped(ws, row, col, value, *, font=BODY, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = WRAP_TOP
    if fill:
        c.fill = fill
    return c


def build_summary_sheet(ws, data):
    add_header_band(ws, data.get("scope", "Whole library"))
    counts = {
        "conflicting_facts": len(data.get("conflicting_facts", [])),
        "outdated_facts": len(data.get("outdated_facts", [])),
        "capability_contradictions": len(data.get("capability_contradictions", [])),
    }
    total = sum(counts.values())

    ws["A6"] = "Generated"; ws["B6"] = datetime.now().strftime("%d %b %Y")
    ws["A7"] = "Library size"; ws["B7"] = data.get("library_size", "n/a")
    ws["A8"] = "Items audited"; ws["B8"] = data.get("items_audited", "n/a")
    ws["A9"] = "Scope"; ws["B9"] = data.get("scope", "Whole library")
    ws["A10"] = "Total contradictions flagged"; ws["B10"] = total
    ws["A11"] = "  - Conflicting facts"; ws["B11"] = counts["conflicting_facts"]
    ws["A12"] = "  - Outdated facts"; ws["B12"] = counts["outdated_facts"]
    ws["A13"] = "  - Capability contradictions"; ws["B13"] = counts["capability_contradictions"]

    for r in range(6, 14):
        ws.cell(row=r, column=1).font = BODY_BOLD
        ws.cell(row=r, column=2).font = BODY

    # Ground truth section
    anchors = data.get("ground_truth_anchors", [])
    if anchors:
        ws["A15"] = "Ground-truth anchors used"
        ws["A15"].font = H1_FONT
        ws["A16"] = "Anchor fact"
        ws["B16"] = "Current truth"
        ws["C16"] = "Source"
        for cidx, col in enumerate(["A", "B", "C"], start=1):
            cell = ws.cell(row=16, column=cidx)
            cell.font = HDR_FONT
            cell.fill = NAVY_FILL
        for i, a in enumerate(anchors, start=17):
            write_wrapped(ws, i, 1, a.get("fact", ""), font=BODY_BOLD)
            write_wrapped(ws, i, 2, a.get("current_truth", ""))
            write_wrapped(ws, i, 3, a.get("source", "user-supplied"))
            ws.row_dimensions[i].height = 30

    # Recommended actions
    actions_row = 17 + len(anchors) + 2
    ws[f"A{actions_row}"] = "Recommended actions"
    ws[f"A{actions_row}"].font = H1_FONT
    actions = [
        ("1. Fix Critical-risk items this week", "Critical items are in content used 10+ times. Every day they sit there, they get pulled into more deals."),
        ("2. Triage Outdated Facts sheet against your ground truth", "These are where a stale answer is putting wrong info in front of buyers right now."),
        ("3. For Capability Contradictions, pick a canonical answer", "Then archive the contradicting versions. A clean library is more valuable than a complete one."),
        ("4. Run this skill again in 90 days", "Library hygiene is recurring work, not one-off. Schedule it."),
    ]
    for i, (a, b) in enumerate(actions, start=actions_row + 1):
        ws[f"A{i}"] = a
        ws[f"A{i}"].font = BODY_BOLD
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        write_wrapped(ws, i, 2, b)
        ws.row_dimensions[i].height = 32

    # Column widths
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 30
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 18


def build_conflicting_facts_sheet(ws, items, scope):
    add_header_band(ws, scope)
    headers = ["Risk", "Topic", "Conflict", "Item A — Question", "Item A — Date", "Item A — Usage", "Item A — Link",
               "Item B — Question", "Item B — Date", "Item B — Usage", "Item B — Link", "Recommendation"]
    row = 6
    for cidx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=cidx, value=h)
        c.font = HDR_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

    items_sorted = sorted(items, key=lambda x: ({"Critical": 0, "High": 1, "Medium": 2, "Verify": 3}.get(x.get("risk"), 9)))

    for idx, item in enumerate(items_sorted, start=row + 1):
        a = item.get("item_a", {})
        b = item.get("item_b", {})
        risk_cell(ws, idx, 1, item.get("risk", "Verify"))
        write_wrapped(ws, idx, 2, item.get("topic", ""), font=BODY_BOLD)
        write_wrapped(ws, idx, 3, item.get("conflict", ""))
        write_wrapped(ws, idx, 4, a.get("question", "")[:300])
        write_wrapped(ws, idx, 5, a.get("date", ""))
        write_wrapped(ws, idx, 6, a.get("usage", ""))
        write_link(ws, idx, 7, a.get("reference_url", ""))
        write_wrapped(ws, idx, 8, b.get("question", "")[:300])
        write_wrapped(ws, idx, 9, b.get("date", ""))
        write_wrapped(ws, idx, 10, b.get("usage", ""))
        write_link(ws, idx, 11, b.get("reference_url", ""))
        write_wrapped(ws, idx, 12, item.get("recommendation", ""))
        ws.row_dimensions[idx].height = 60

    widths = [12, 22, 40, 40, 14, 12, 22, 40, 14, 12, 22, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A7"
    if items_sorted:
        ws.auto_filter.ref = f"A6:L{6 + len(items_sorted)}"


def build_outdated_facts_sheet(ws, items, scope):
    add_header_band(ws, scope)
    headers = ["Risk", "Topic", "Stale Claim in Library", "Anchor (current truth)",
               "Question", "Source File", "Date", "Usage", "Link", "Recommendation"]
    row = 6
    for cidx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=cidx, value=h)
        c.font = HDR_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

    items_sorted = sorted(items, key=lambda x: ({"Critical": 0, "High": 1, "Medium": 2, "Verify": 3}.get(x.get("risk"), 9)))

    for idx, item in enumerate(items_sorted, start=row + 1):
        it = item.get("item", {})
        risk_cell(ws, idx, 1, item.get("risk", "Verify"))
        write_wrapped(ws, idx, 2, item.get("topic", ""), font=BODY_BOLD)
        write_wrapped(ws, idx, 3, item.get("stale_claim", ""))
        write_wrapped(ws, idx, 4, item.get("anchor_fact", ""))
        write_wrapped(ws, idx, 5, it.get("question", "")[:300])
        write_wrapped(ws, idx, 6, it.get("file", ""))
        write_wrapped(ws, idx, 7, it.get("date", ""))
        write_wrapped(ws, idx, 8, it.get("usage", ""))
        write_link(ws, idx, 9, it.get("reference_url", ""))
        write_wrapped(ws, idx, 10, item.get("recommendation", ""))
        ws.row_dimensions[idx].height = 55

    widths = [12, 22, 44, 36, 44, 28, 12, 11, 22, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A7"
    if items_sorted:
        ws.auto_filter.ref = f"A6:J{6 + len(items_sorted)}"


def build_capability_sheet(ws, items, scope):
    add_header_band(ws, scope)
    headers = ["Risk", "Capability", "Negative Claim — Question", "Negative — File", "Negative — Date",
               "Negative — Usage", "Negative — Link", "Positive Claim — Question", "Positive — File",
               "Positive — Date", "Positive — Usage", "Positive — Link", "Recommendation"]
    row = 6
    for cidx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=cidx, value=h)
        c.font = HDR_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

    items_sorted = sorted(items, key=lambda x: ({"Critical": 0, "High": 1, "Medium": 2, "Verify": 3}.get(x.get("risk"), 9)))

    for idx, item in enumerate(items_sorted, start=row + 1):
        neg = item.get("negative_item", {})
        pos = item.get("positive_item", {})
        risk_cell(ws, idx, 1, item.get("risk", "Verify"))
        write_wrapped(ws, idx, 2, item.get("capability", ""), font=BODY_BOLD)
        write_wrapped(ws, idx, 3, neg.get("question", "")[:300])
        write_wrapped(ws, idx, 4, neg.get("file", ""))
        write_wrapped(ws, idx, 5, neg.get("date", ""))
        write_wrapped(ws, idx, 6, neg.get("usage", ""))
        write_link(ws, idx, 7, neg.get("reference_url", ""))
        write_wrapped(ws, idx, 8, pos.get("question", "")[:300])
        write_wrapped(ws, idx, 9, pos.get("file", ""))
        write_wrapped(ws, idx, 10, pos.get("date", ""))
        write_wrapped(ws, idx, 11, pos.get("usage", ""))
        write_link(ws, idx, 12, pos.get("reference_url", ""))
        write_wrapped(ws, idx, 13, item.get("recommendation", ""))
        ws.row_dimensions[idx].height = 60

    widths = [12, 24, 36, 24, 12, 11, 22, 36, 24, 12, 11, 22, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A7"
    if items_sorted:
        ws.auto_filter.ref = f"A6:M{6 + len(items_sorted)}"


def build_workbook(data: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    build_summary_sheet(ws, data)

    cf = wb.create_sheet("Conflicting Facts")
    build_conflicting_facts_sheet(cf, data.get("conflicting_facts", []), data.get("scope", ""))

    of = wb.create_sheet("Outdated Facts")
    build_outdated_facts_sheet(of, data.get("outdated_facts", []), data.get("scope", ""))

    cc = wb.create_sheet("Capability Contradictions")
    build_capability_sheet(cc, data.get("capability_contradictions", []), data.get("scope", ""))

    wb.save(output_path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, help="Path to findings JSON")
    p.add_argument("--output", required=True, help="Path to output .xlsx")
    args = p.parse_args()

    with open(args.input) as f:
        data = json.load(f)
    build_workbook(data, args.output)
    print(f"Saved: {args.output}")
    print(f"  Conflicting facts:        {len(data.get('conflicting_facts', []))}")
    print(f"  Outdated facts:           {len(data.get('outdated_facts', []))}")
    print(f"  Capability contradictions: {len(data.get('capability_contradictions', []))}")


if __name__ == "__main__":
    main()
